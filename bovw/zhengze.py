import re
import xlwt
import xlwings as xw
import numpy as np
import pandas as pd
from openpyxl import load_workbook

for number in range(0,20):
    sum1 = 0
    sum2 = 0
    sum3 = 0
    sum4 = 0
    sum5 = 0
    sum6 = 0
    sum7 = 0
    sum8 = 0
    sum9 = 0
    sum10 = 0
    sum11 = 0
    sum12 = 0
    sum13 = 0
    sum14 = 0
    sum15 = 0
    sum16 = 0
    sum17 = 0
    sum18 = 0
    sum19 = 0
    sum20 = 0
    sum21 = 0
    sum22 = 0
    sum23 = 0
    sum24 = 0
    sum25 = 0
    sum1v = []
    sum2v = []
    sum3v = []
    sum4v = []
    sum5v = []
    sum6v = []
    sum7v = []
    sum8v = []
    sum9v = []
    sum10v = []
    sum11v = []
    sum12v = []
    sum13v = []
    sum14v = []
    sum15v = []
    sum16v = []
    sum17v = []
    sum18v = []
    sum19v = []
    sum20v = []
    sum21v = []
    sum22v = []
    sum23v = []
    sum24v = []
    sum25v = []

    sum_a = []
    #number = int(input("input number: "))
    sum1min = 1
    sum1max = 0
    sum2min = 1
    sum2max = 0
    sum3min = 1
    sum3max = 0
    sum4min = 1
    sum4max = 0
    sum5min = 1
    sum5max = 0
    sum6min = 1
    sum6max = 0
    sum7min = 1
    sum7max = 0
    sum8min = 1
    sum8max = 0
    sum9min = 1
    sum9max = 0
    sum10min = 1
    sum10max = 0
    sum11min = 1
    sum11max = 0
    sum12min = 1
    sum12max = 0
    sum13min = 1
    sum13max = 0
    sum14min = 1
    sum14max = 0
    sum15min = 1
    sum15max = 0
    sum16min = 1
    sum16max = 0
    sum17min = 1
    sum17max = 0
    sum18min = 1
    sum18max = 0
    sum19min = 1
    sum19max = 0
    sum20min = 1
    sum20max = 0
    sum21min = 1
    sum21max = 0
    sum22min = 1
    sum22max = 0
    sum23min = 1
    sum23max = 0
    sum24min = 1
    sum24max = 0
    sum25min = 1
    sum25max = 0
    for i in range(1,11):

        f = open(r"I:\image_classification_using_sift\manual_cross\svmc5\\" + "cross" + str(i)+"surf_svm_5_record.txt",'r')
        #f = open('I:\image_classification_using_sift\manual_cross\cross1record.txt','a', encoding='utf-8')
        s=f.read()

        Babcock_Tissue_Forceps_6 = re.findall( r'6_Babcock_Tissue_Forceps...........................(.*?)\ ', s)
        Mayo_Needle_Holder_6 = re.findall( r'6_Mayo_Needle_Holder...........................(.*?)\ ', s)
        Metzenbaum_Scissors_7 = re.findall( r'7_Metzenbaum_Scissors...........................(.*?)\ ', s)
        Microvascular_Needle_Holder_7 = re.findall( r'7_Microvascular_Needle_Holder...........................(.*?)\ ', s)
        Babcock_Tissue_Forceps_8 = re.findall( r'8_Babcock_Tissue_Forceps...........................(.*?)\ ', s)
        Mayo_Needle_Holder_8 = re.findall( r'8_Mayo_Needle_Holder...........................(.*?)\ ', s)
        Microvascular_Needle_Holder_8 = re.findall( r'8_Microvascular_Needle_Holder...........................(.*?)\ ', s)
        DeBakey_Dissector_9 = re.findall( r'9_DeBakey_Dissector...........................(.*?)\ ', s)
        DeBakey_Needle_Holder_9 = re.findall( r'9_DeBakey_Needle_Holder...........................(.*?)\ ', s)
        Metzenbaum_Scissors_9 = re.findall( r'9_Metzenbaum_Scissors...........................(.*?)\ ', s)
        Allis_Tissue_Forceps = re.findall( r'Allis_Tissue_Forceps...........................(.*?)\ ', s)
        Ball___Socket_Towel_Clips = re.findall( r'Ball___Socket_Towel_Clips...........................(.*?)\ ', s)
        Bonneys_Non_Toothed_Dissector = re.findall( r'Bonneys_Non_Toothed_Dissector...........................(.*?)\ ', s)
        Bonneys_Toothed_Dissector = re.findall( r'Bonneys_Toothed_Dissector...........................(.*?)\ ', s)
        Crile_Artery_Forceps = re.findall( r'Crile_Artery_Forceps...........................(.*?)\ ', s)
        Curved_Mayo_Scissors = re.findall( r'Curved_Mayo_Scissors...........................(.*?)\ ', s)
        Dressing_Scissors = re.findall( r'Dressing_Scissors...........................(.*?)\ ', s)
        Gillies_Toothed_Dissector  = re.findall( r'Gillies_Toothed_Dissector...........................(.*?)\ ', s)
        Lahey_Forceps = re.findall( r'Lahey_Forceps...........................(.*?)\ ', s)
        Littlewood_Tissue_Forceps = re.findall( r'Littlewood_Tissue_Forceps...........................(.*?)\ ', s)
        Mayo_Artery_Forceps = re.findall( r'Mayo_Artery_Forceps...........................(.*?)\ ', s)
        No3_BP_Handles = re.findall( r'No3_BP_Handles...........................(.*?)\ ', s)
        No4_BP_Handles = re.findall( r'No4_BP_Handles...........................(.*?)\ ', s)
        No7_BP_Handles = re.findall( r'No7_BP_Handles...........................(.*?)\ ', s)
        Sponge_Forceps = re.findall( r'Sponge_Forceps...........................(.*?)\ ', s)

        #sum = float(Babcock_Tissue_Forceps_6[number]) + sum
        sum1 = float(Babcock_Tissue_Forceps_6[number]) + sum1
        sum1v.append(float(Babcock_Tissue_Forceps_6[number]))
        if sum1min > float(Babcock_Tissue_Forceps_6[number]):
            sum1min = float(Babcock_Tissue_Forceps_6[number])
        if sum1max < float(Babcock_Tissue_Forceps_6[number]):
            sum1max = float(Babcock_Tissue_Forceps_6[number])

        sum2 = float(Mayo_Needle_Holder_6[number]) + sum2
        sum2v.append(float(Mayo_Needle_Holder_6[number]))
        if sum2min > float(Mayo_Needle_Holder_6[number]):
            sum2min = float(Mayo_Needle_Holder_6[number])
        if sum2max < float(Mayo_Needle_Holder_6[number]):
            sum2max = float(Mayo_Needle_Holder_6[number])

        sum3 = float(Metzenbaum_Scissors_7[number]) + sum3
        sum3v.append(float(Metzenbaum_Scissors_7[number]))
        if sum3min > float(Metzenbaum_Scissors_7[number]):
            sum3min = float(Metzenbaum_Scissors_7[number])
        if sum3max < float(Metzenbaum_Scissors_7[number]):
            sum3max = float(Metzenbaum_Scissors_7[number])

        sum4 = float(Microvascular_Needle_Holder_7[number]) + sum4
        sum4v.append(float(Microvascular_Needle_Holder_7[number]))
        if sum4min > float(Microvascular_Needle_Holder_7[number]):
            sum4min = float(Microvascular_Needle_Holder_7[number])
        if sum4max < float(Microvascular_Needle_Holder_7[number]):
            sum4max = float(Microvascular_Needle_Holder_7[number])


        sum5 = float(Babcock_Tissue_Forceps_8[number]) + sum5
        sum5v.append(float(Babcock_Tissue_Forceps_8[number]))
        if sum5min > float(Babcock_Tissue_Forceps_8[number]):
            sum5min = float(Babcock_Tissue_Forceps_8[number])
        if sum5max < float(Babcock_Tissue_Forceps_8[number]):
            sum5max = float(Babcock_Tissue_Forceps_8[number])

        sum6 = float(Mayo_Needle_Holder_8[number]) + sum6
        sum6v.append(float(Mayo_Needle_Holder_8[number]))
        if sum6min > float(Mayo_Needle_Holder_8[number]):
            sum6min = float(Mayo_Needle_Holder_8[number])
        if sum6max < float(Mayo_Needle_Holder_8[number]):
            sum6max = float(Mayo_Needle_Holder_8[number])

        sum7 = float(Microvascular_Needle_Holder_8[number]) + sum7
        sum7v.append(float(Microvascular_Needle_Holder_8[number]))
        if sum7min > float(Microvascular_Needle_Holder_8[number]):
            sum7min = float(Microvascular_Needle_Holder_8[number])
        if sum7max < float(Microvascular_Needle_Holder_8[number]):
            sum7max = float(Microvascular_Needle_Holder_8[number])

        sum8 = float(DeBakey_Dissector_9[number]) + sum8
        sum8v.append(float(DeBakey_Dissector_9[number]))
        if sum8min > float(DeBakey_Dissector_9[number]):
            sum8min = float(DeBakey_Dissector_9[number])
        if sum8max < float(DeBakey_Dissector_9[number]):
            sum8max = float(DeBakey_Dissector_9[number])

        sum9 = float(DeBakey_Needle_Holder_9[number]) + sum9
        sum9v.append(float(DeBakey_Needle_Holder_9[number]))
        if sum9min > float(DeBakey_Needle_Holder_9[number]):
            sum9min = float(DeBakey_Needle_Holder_9[number])
        if sum9max < float(DeBakey_Needle_Holder_9[number]):
            sum9max = float(DeBakey_Needle_Holder_9[number])

        sum10 = float(Metzenbaum_Scissors_9[number]) + sum10
        sum10v.append(float(Metzenbaum_Scissors_9[number]))
        if sum10min > float(Metzenbaum_Scissors_9[number]):
            sum10min = float(Metzenbaum_Scissors_9[number])
        if sum10max < float(Metzenbaum_Scissors_9[number]):
            sum10max = float(Metzenbaum_Scissors_9[number])

        sum11 = float(Allis_Tissue_Forceps[number]) + sum11
        sum11v.append(float(Allis_Tissue_Forceps[number]))
        if sum11min > float(Allis_Tissue_Forceps[number]):
            sum11min = float(Allis_Tissue_Forceps[number])
        if sum11max < float(Allis_Tissue_Forceps[number]):
            sum11max = float(Allis_Tissue_Forceps[number])

        sum12 = float(Ball___Socket_Towel_Clips[number]) + sum12
        sum12v.append(float(Ball___Socket_Towel_Clips[number]))
        if sum12min > float(Ball___Socket_Towel_Clips[number]):
            sum12min = float(Ball___Socket_Towel_Clips[number])
        if sum12max < float(Ball___Socket_Towel_Clips[number]):
            sum12max = float(Ball___Socket_Towel_Clips[number])

        sum13 = float(Bonneys_Non_Toothed_Dissector[number]) + sum13
        sum13v.append(float(Bonneys_Non_Toothed_Dissector[number]))
        if sum13min > float(Bonneys_Non_Toothed_Dissector[number]):
            sum13min = float(Bonneys_Non_Toothed_Dissector[number])
        if sum13max < float(Bonneys_Non_Toothed_Dissector[number]):
            sum13max = float(Bonneys_Non_Toothed_Dissector[number])

        sum14 = float(Bonneys_Toothed_Dissector[number]) + sum14
        sum14v.append(float(Bonneys_Toothed_Dissector[number]))
        if sum14min > float(Bonneys_Toothed_Dissector[number]):
            sum14min = float(Bonneys_Toothed_Dissector[number])
        if sum14max < float(Bonneys_Toothed_Dissector[number]):
            sum14max = float(Bonneys_Toothed_Dissector[number])

        sum15 = float(Crile_Artery_Forceps[number]) + sum15
        sum15v.append(float(Crile_Artery_Forceps[number]))
        if sum15min > float(Crile_Artery_Forceps[number]):
            sum15min = float(Crile_Artery_Forceps[number])
        if sum15max < float(Crile_Artery_Forceps[number]):
            sum15max = float(Crile_Artery_Forceps[number])

        sum16 = float(Curved_Mayo_Scissors[number]) + sum16
        sum16v.append(float(Curved_Mayo_Scissors[number]))
        if sum16min > float(Curved_Mayo_Scissors[number]):
            sum16min = float(Curved_Mayo_Scissors[number])
        if sum16max < float(Curved_Mayo_Scissors[number]):
            sum16max = float(Curved_Mayo_Scissors[number])

        sum17 = float(Dressing_Scissors[number]) + sum17
        sum17v.append(float(Dressing_Scissors[number]))
        if sum17min > float(Dressing_Scissors[number]):
            sum17min = float(Dressing_Scissors[number])
        if sum17max < float(Dressing_Scissors[number]):
            sum17max = float(Dressing_Scissors[number])

        sum18 = float(Gillies_Toothed_Dissector[number]) + sum18
        sum18v.append(float(Gillies_Toothed_Dissector[number]))
        if sum18min > float(Gillies_Toothed_Dissector[number]):
            sum18min = float(Gillies_Toothed_Dissector[number])
        if sum18max < float(Gillies_Toothed_Dissector[number]):
            sum18max = float(Gillies_Toothed_Dissector[number])

        sum19 = float(Lahey_Forceps[number]) + sum19
        sum19v.append(float(Lahey_Forceps[number]))
        if sum19min > float(Lahey_Forceps[number]):
            sum19min = float(Lahey_Forceps[number])
        if sum19max < float(Lahey_Forceps[number]):
            sum19max = float(Lahey_Forceps[number])

        sum20 = float(Littlewood_Tissue_Forceps[number]) + sum20
        sum20v.append(float(Littlewood_Tissue_Forceps[number]))
        if sum20min > float(Littlewood_Tissue_Forceps[number]):
            sum20min = float(Littlewood_Tissue_Forceps[number])
        if sum20max < float(Littlewood_Tissue_Forceps[number]):
            sum20max = float(Littlewood_Tissue_Forceps[number])

        sum21 = float(Mayo_Artery_Forceps[number]) + sum21
        sum21v.append(float(Mayo_Artery_Forceps[number]))
        if sum21min > float(Mayo_Artery_Forceps[number]):
            sum21min = float(Mayo_Artery_Forceps[number])
        if sum21max < float(Mayo_Artery_Forceps[number]):
            sum21max = float(Mayo_Artery_Forceps[number])

        sum22 = float(No3_BP_Handles[number]) + sum22
        sum22v.append(float(No3_BP_Handles[number]))
        if sum22min > float(No3_BP_Handles[number]):
            sum22min = float(No3_BP_Handles[number])
        if sum22max < float(No3_BP_Handles[number]):
            sum22max = float(No3_BP_Handles[number])

        sum23 = float(No4_BP_Handles[number]) + sum23
        sum23v.append(float(No4_BP_Handles[number]))
        if sum23min > float(No4_BP_Handles[number]):
            sum23min = float(No4_BP_Handles[number])
        if sum23max < float(No4_BP_Handles[number]):
            sum23max = float(No4_BP_Handles[number])

        sum24 = float(No7_BP_Handles[number]) + sum24
        sum24v.append(float(No7_BP_Handles[number]))
        if sum24min > float(No7_BP_Handles[number]):
            sum24min = float(No7_BP_Handles[number])
        if sum24max < float(No7_BP_Handles[number]):
            sum24max = float(No7_BP_Handles[number])

        sum25 = float(Sponge_Forceps[number]) + sum25
        sum25v.append(float(Sponge_Forceps[number]))
        if sum25min > float(Sponge_Forceps[number]):
            sum25min = float(Sponge_Forceps[number])
        if sum25max < float(Sponge_Forceps[number]):
            sum25max = float(Sponge_Forceps[number])

    
    # sum1 = str(round(sum1/10,3)) + "(+ " + str(round(sum1max-round(sum1/10,3),3)) +" - "+str(round(round(sum1/10,3)-sum1min,3))+")"
    # sum2 = str(round(sum2/10,3)) + "(+ " + str(round(sum2max-round(sum2/10,3),3)) +" - "+str(round(round(sum2/10,3)-sum2min,3))+")"
    # sum3 = str(round(sum3/10,3)) + "(+ " + str(round(sum3max-round(sum3/10,3),3)) +" - "+str(round(round(sum3/10,3)-sum3min,3))+")"
    # sum4 = str(round(sum4/10,3)) + "(+ " + str(round(sum4max-round(sum4/10,3),3)) +" - "+str(round(round(sum4/10,3)-sum4min,3))+")"
    # sum5 = str(round(sum5/10,3)) + "(+ " + str(round(sum5max-round(sum5/10,3),3)) +" - "+str(round(round(sum5/10,3)-sum5min,3))+")"
    # sum6 = str(round(sum6/10,3)) + "(+ " + str(round(sum6max-round(sum6/10,3),3)) +" - "+str(round(round(sum6/10,3)-sum6min,3))+")"
    # sum7 = str(round(sum7/10,3)) + "(+ " + str(round(sum7max-round(sum7/10,3),3)) +" - "+str(round(round(sum7/10,3)-sum7min,3))+")"
    # sum8 = str(round(sum8/10,3)) + "(+ " + str(round(sum8max-round(sum8/10,3),3)) +" - "+str(round(round(sum8/10,3)-sum8min,3))+")"
    # sum9 = str(round(sum9/10,3)) + "(+ " + str(round(sum9max-round(sum9/10,3),3)) +" - "+str(round(round(sum9/10,3)-sum9min,3))+")"
    # sum10 = str(round(sum10/10,3)) + "(+ " + str(round(sum10max-round(sum10/10,3),3)) +" - "+str(round(round(sum10/10,3)-sum10min,3))+")"
    # sum11 = str(round(sum11/10,3)) + "(+ " + str(round(sum11max-round(sum11/10,3),3)) +" - "+str(round(round(sum11/10,3)-sum11min,3))+")"
    # sum12 = str(round(sum12/10,3)) + "(+ " + str(round(sum12max-round(sum12/10,3),3)) +" - "+str(round(round(sum12/10,3)-sum12min,3))+")"
    # sum13 = str(round(sum13/10,3)) + "(+ " + str(round(sum13max-round(sum13/10,3),3)) +" - "+str(round(round(sum13/10,3)-sum13min,3))+")"
    # sum14 = str(round(sum14/10,3)) + "(+ " + str(round(sum14max-round(sum14/10,3),3)) +" - "+str(round(round(sum14/10,3)-sum14min,3))+")"
    # sum15 = str(round(sum15/10,3)) + "(+ " + str(round(sum15max-round(sum15/10,3),3)) +" - "+str(round(round(sum15/10,3)-sum15min,3))+")"
    # sum16 = str(round(sum16/10,3)) + "(+ " + str(round(sum16max-round(sum16/10,3),3)) +" - "+str(round(round(sum16/10,3)-sum16min,3))+")"
    # sum17 = str(round(sum17/10,3)) + "(+ " + str(round(sum17max-round(sum17/10,3),3)) +" - "+str(round(round(sum17/10,3)-sum17min,3))+")"
    # sum18 = str(round(sum18/10,3)) + "(+ " + str(round(sum18max-round(sum18/10,3),3)) +" - "+str(round(round(sum18/10,3)-sum18min,3))+")"
    # sum19 = str(round(sum19/10,3)) + "(+ " + str(round(sum19max-round(sum19/10,3),3)) +" - "+str(round(round(sum19/10,3)-sum19min,3))+")"
    # sum20 = str(round(sum20/10,3)) + "(+ " + str(round(sum20max-round(sum20/10,3),3)) +" - "+str(round(round(sum20/10,3)-sum20min,3))+")"
    # sum21 = str(round(sum21/10,3)) + "(+ " + str(round(sum21max-round(sum21/10,3),3)) +" - "+str(round(round(sum21/10,3)-sum21min,3))+")"
    # sum22 = str(round(sum22/10,3)) + "(+ " + str(round(sum22max-round(sum22/10,3),3)) +" - "+str(round(round(sum22/10,3)-sum22min,3))+")"
    # sum23 = str(round(sum23/10,3)) + "(+ " + str(round(sum23max-round(sum23/10,3),3)) +" - "+str(round(round(sum23/10,3)-sum23min,3))+")"
    # sum24 = str(round(sum24/10,3)) + "(+ " + str(round(sum24max-round(sum24/10,3),3)) +" - "+str(round(round(sum24/10,3)-sum24min,3))+")"
    # sum25 = str(round(sum25/10,3)) + "(+ " + str(round(sum25max-round(sum25/10,3),3)) +" - "+str(round(round(sum25/10,3)-sum25min,3))+")"


    sum1 = str(round(sum1/10,3)) + "( " + str(round(np.std(sum1v),3)) + ")"
    sum2 = str(round(sum2/10,3)) + "( " + str(round(np.std(sum2v),3))+")"
    sum3 = str(round(sum3/10,3)) + "( " + str(round(np.std(sum3v),3))+")"
    sum4 = str(round(sum4/10,3)) + "( " + str(round(np.std(sum4v),3))+")"
    sum5 = str(round(sum5/10,3)) + "( " + str(round(np.std(sum5v),3))+")"
    sum6 = str(round(sum6/10,3)) + "( " + str(round(np.std(sum6v),3))+")"
    sum7 = str(round(sum7/10,3)) + "( " + str(round(np.std(sum7v),3))+")"
    sum8 = str(round(sum8/10,3)) + "( " + str(round(np.std(sum8v),3))+")"
    sum9 = str(round(sum9/10,3)) + "( " + str(round(np.std(sum9v),3))+")"
    sum10 = str(round(sum10/10,3)) + "( " + str(round(np.std(sum10v),3))+")"
    sum11 = str(round(sum11/10,3)) + "( " + str(round(np.std(sum11v),3))+")"
    sum12 = str(round(sum12/10,3)) + "( " + str(round(np.std(sum12v),3))+")"
    sum13 = str(round(sum13/10,3)) + "( " + str(round(np.std(sum13v),3))+")"
    sum14 = str(round(sum14/10,3)) + "( " + str(round(np.std(sum14v),3))+")"
    sum15 = str(round(sum15/10,3)) + "( " + str(round(np.std(sum15v),3))+")"
    sum16 = str(round(sum16/10,3)) + "( " + str(round(np.std(sum16v),3))+")"
    sum17 = str(round(sum17/10,3)) + "( " + str(round(np.std(sum17v),3))+")"
    sum18 = str(round(sum18/10,3)) + "( " + str(round(np.std(sum18v),3))+")"
    sum19 = str(round(sum19/10,3)) + "( " + str(round(np.std(sum19v),3))+")"
    sum20 = str(round(sum20/10,3)) + "( " + str(round(np.std(sum20v),3))+")"
    sum21 = str(round(sum21/10,3)) + "( " + str(round(np.std(sum21v),3))+")"
    sum22 = str(round(sum22/10,3)) + "( " + str(round(np.std(sum22v),3))+")"
    sum23 = str(round(sum23/10,3)) + "( " + str(round(np.std(sum23v),3))+")"
    sum24 = str(round(sum24/10,3)) + "( " + str(round(np.std(sum24v),3))+")"
    sum25 = str(round(sum25/10,3)) + "( " + str(round(np.std(sum25v),3))+")"

    def trans(arr):
        arr = np.array(arr)


    # trans(sum1v)
    # trans(sum2v)
    # trans(sum3v)
    # trans(sum4v)
    # trans(sum5v)
    # trans(sum6v)
    # trans(sum7v)
    # trans(sum8v)
    # trans(sum9v)
    # trans(sum10v)
    # trans(sum11v)
    # trans(sum12v)
    # trans(sum13v)
    # trans(sum14v)
    # trans(sum15v)
    # trans(sum16v)
    # trans(sum17v)
    # trans(sum18v)
    # trans(sum19v)
    # trans(sum20v)
    # trans(sum21v)
    # trans(sum22v)
    # trans(sum23v)
    # trans(sum24v)
    # trans(sum25v)

    average_f1 = round((round(sum(sum1v)/10,3)+round(sum(sum2v)/10,3)+round(sum(sum3v)/10,3)+round(sum(sum4v)/10,3)+round(sum(sum5v)/10,3)+round(sum(sum6v)/10,3)+round(sum(sum7v)/10,3)+round(sum(sum8v)/10,3)+round(sum(sum9v)/10,3)+round(sum(sum10v)/10,3)+round(sum(sum11v)/10,3)+round(sum(sum12v)/10,3)+round(sum(sum13v)/10,3)+round(sum(sum14v)/10,3)+round(sum(sum15v)/10,3)+round(sum(sum16v)/10,3)+round(sum(sum17v)/10,3)+round(sum(sum18v)/10,3)+round(sum(sum19v)/10,3)+round(sum(sum20v)/10,3)+round(sum(sum21v)/10,3)+round(sum(sum22v)/10,3)+round(sum(sum23v)/10,3)+round(sum(sum24v)/10,3)+round(sum(sum25v)/10,3))/25,3)

    sum_a.append(sum1)   
    sum_a.append(sum2) 
    sum_a.append(sum3) 
    sum_a.append(sum4) 
    sum_a.append(sum5) 
    sum_a.append(sum6) 
    sum_a.append(sum7) 
    sum_a.append(sum8) 
    sum_a.append(sum9) 
    sum_a.append(sum10) 
    sum_a.append(sum11) 
    sum_a.append(sum12) 
    sum_a.append(sum13) 
    sum_a.append(sum14) 
    sum_a.append(sum15) 
    sum_a.append(sum16) 
    sum_a.append(sum17) 
    sum_a.append(sum18) 
    sum_a.append(sum19) 
    sum_a.append(sum20) 
    sum_a.append(sum21) 
    sum_a.append(sum22) 
    sum_a.append(sum23) 
    sum_a.append(sum24) 
    sum_a.append(sum25) 
    sum_a.append(str(average_f1))

    #sum = [round(i,3) for i in sum]

    # book = load_workbook('A.xlsx')
    # data = pd.DataFrame(sum)
    # #writer = pd.ExcelWriter('A.xlsx')       # 写入Excel文件
    # data.to_excel(book, 'page_1', float_format='%.5f')        # ‘page_1’是写入excel的sheet名
    # writer.save()
    # writer.close()

    data = pd.DataFrame(sum_a)
    book = load_workbook('A.xlsx')
    with pd.ExcelWriter('A.xlsx') as E:
        E.book = book
        E.sheets = dict((ws.title, ws) for ws in book.worksheets)
        data.to_excel(E, str(number), float_format='%.5f')
